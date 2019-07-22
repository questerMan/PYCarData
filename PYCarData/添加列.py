#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 17 14:32:19 2019

@author: lixin
"""

from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet() #默认插在工作簿末尾

ws2 = wb.create_sheet(0) # 插入在工作簿的第一个位置

ws.title = "New Title"

ws.sheet_properties.tabColor = "1072BA"

ws3 = wb["New Title"]

ws4 = wb.get_sheet_by_name("New Title")

print(ws is ws3 is ws4)

print(wb.get_sheet_names())